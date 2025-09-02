# db_operations.py

import os
import re
import pandas as pd
import mysql.connector
from mysql.connector import Error
from datetime import datetime
import calendar
import config
from tqdm import tqdm
import openpyxl


def create_connection(host_name, user_name, user_password, db_name=None):
    try:
        connection = mysql.connector.connect(
            host=host_name, user=user_name, passwd=user_password, database=db_name if db_name else None
        )
        print(f"MySQL connection successful ({'DB: ' + db_name if db_name else 'server'})")
        return connection
    except Error as e:
        print(f"The error '{e}' occurred")
        return None


def create_database(connection, db_name):
    cursor = connection.cursor()
    try:
        cursor.execute(f"CREATE DATABASE IF NOT EXISTS {db_name}")
        print(f"Database '{db_name}' created or already exists.")
    except Error as e:
        print(f"Error creating database {db_name}: {e}")


def create_pmr_table(connection):
    cursor = connection.cursor()
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS {config.PMR_TABLE} (
            PROJECT_ID VARCHAR(255) PRIMARY KEY,
            PGM_MANAGER_NAME VARCHAR(255),
            PGM_MANAGER_EMAIL VARCHAR(255)
        );
    """)
    print(f"Table '{config.PMR_TABLE}' is ready in the global PMR database.")


def create_account_tables(connection):
    cursor = connection.cursor()
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS {config.REGIONAL_TABLE} (
            id INT AUTO_INCREMENT PRIMARY KEY, fiscal_year VARCHAR(10),
            EMPLID VARCHAR(255), CURRENT_WORK_LOCATION VARCHAR(255), 
            PROJECT_ID VARCHAR(255), PROJECT_DESCRIPTION TEXT, 
            PROJECT_TYPE VARCHAR(255), CONTRACT_TYPE VARCHAR(255), 
            CUST_NAME VARCHAR(255), RUS_STATUS VARCHAR(255), 
            TOTAL_HOURS DECIMAL(10, 2), Month DATE
        );
    """)

    # --- MODIFIED: Added optional ER_NIC_SUM column ---
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS {config.SALARY_TABLE} (
            id INT AUTO_INCREMENT PRIMARY KEY, fiscal_year VARCHAR(10),
            EMPLID VARCHAR(255), 
            MONTH DATE, 
            GROSS_PAY DECIMAL(10, 2),
            ER_NIC_SUM DECIMAL(10, 2)
        );
    """)

    # --- MODIFIED: Added ER_NIC_SUM column ---
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS {config.CONSOLIDATION_TABLE} (
            id INT AUTO_INCREMENT PRIMARY KEY, fiscal_year VARCHAR(10),
            EMPLID VARCHAR(255), Month DATE, GROSS_PAY DECIMAL(10, 2), 
            ER_NIC_SUM DECIMAL(10, 2),
            DESIGNATION VARCHAR(255), BAND VARCHAR(255), `FUNCTION` VARCHAR(255),
            CURRENT_WORK_LOCATION VARCHAR(255), PROJECT_ID VARCHAR(255), 
            PROJECT_DESCRIPTION TEXT, PROJECT_TYPE VARCHAR(255), 
            CONTRACT_TYPE VARCHAR(255), CUST_NAME VARCHAR(255),
            PGM_MANAGER_NAME VARCHAR(255), PGM_MANAGER_EMAIL VARCHAR(255),
            UNIQUE KEY `unique_employee_month_project_year` (`EMPLID`(100),`Month`,`PROJECT_ID`(100),`fiscal_year`)
        );
    """)
    print("All account-specific tables are ready.")


# ▼▼▼ THIS FUNCTION HAS BEEN UPDATED ▼▼▼
def create_abd_table(connection):
    """
    Creates the associate_base_data table with an auto-incrementing ID
    to allow for multiple records per employee.
    """
    cursor = connection.cursor()
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS {config.ABD_TABLE_NAME} (
            id INT AUTO_INCREMENT PRIMARY KEY,
            EMPLID VARCHAR(255),
            `Function` VARCHAR(255),
            Designation VARCHAR(255),
            BAND VARCHAR(255),
            PROGRAM_MANAGER_NAME VARCHAR(255),
            CURRENT_LOCATION_DESCRIPTION VARCHAR(255),
            PROJECT_ID VARCHAR(255),
            PROJECT_DESCRIPTION TEXT
        );
    """)
    print(f"Table '{config.ABD_TABLE_NAME}' is ready in the global ABD database.")


# ▲▲▲ END OF UPDATED SECTION ▲▲▲


# ▼▼▼ THIS FUNCTION HAS BEEN COMPLETELY REWRITTEN ▼▼▼
def import_abd_data(connection, abd_folder_path):
    """
    Finds all ABD Excel files, clears the ABD table, and loads all rows
    from each file's 'base' or 'data' sheet directly into the database.
    """
    cursor = connection.cursor()

    # Clear the table once before loading any new data
    print("Clearing the existing Associate Base Data table...")
    cursor.execute(f"TRUNCATE TABLE {config.ABD_TABLE_NAME}")
    print("Table cleared.")

    # Regex to find files like 'ABD_Mar-24.xlsx'
    abd_file_pattern = re.compile(r"ABD_[A-Za-z]{3}-\d{2}\.xlsx?$")
    abd_files = [f for f in os.listdir(abd_folder_path) if abd_file_pattern.match(f)]
    total_files = len(abd_files)

    if not abd_files:
        print("No ABD files found matching the pattern 'ABD_{mmm}-{yy}.xlsx'.")
        return

    print(f"Found {total_files} ABD file(s) to process.")

    # Process each file individually
    for i, file_name in enumerate(sorted(abd_files), 1):
        print(f"\n--- [{i}/{total_files}] Processing file: {file_name} ---")
        file_path = os.path.join(abd_folder_path, file_name)

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            if not workbook.sheetnames:
                print(f"  -> ERROR: File '{file_name}' contains no sheets. Skipping.")
                continue

            # Find the target sheet ('base', 'data', or fallback to first)
            target_sheet_obj = None
            clean_sheet_names = {s.strip().lower(): s for s in workbook.sheetnames}
            if 'base' in clean_sheet_names:
                target_sheet_obj = workbook[clean_sheet_names['base']]
            elif 'data' in clean_sheet_names:
                target_sheet_obj = workbook[clean_sheet_names['data']]
            else:
                target_sheet_obj = workbook.worksheets[0]
                print(
                    f"  -> WARNING: Could not find 'base' or 'data' sheet. Using first sheet: '{target_sheet_obj.title}'")

            header = [cell.value for cell in target_sheet_obj[1]]
            header_upper = [str(h).strip().upper() for h in header]

            # Map Excel columns to database columns based on config
            col_map = {}
            for excel_key, db_col_name in config.ABD_COLUMN_MAP.items():
                sanitized_key = excel_key.replace('_', '').replace(' ', '')
                for idx, actual_header in enumerate(header_upper):
                    sanitized_header = actual_header.replace('_', '').replace(' ', '')
                    if sanitized_header.startswith(sanitized_key):
                        if db_col_name not in col_map:
                            col_map[db_col_name] = idx
                            break

            if 'EMPLID' not in col_map:
                print(
                    f"  -> ERROR: Required column 'EMPLID' not found in sheet '{target_sheet_obj.title}'. Skipping file.")
                continue

            # Prepare SQL statement
            db_cols = list(col_map.keys())
            col_str = ", ".join([f"`{c}`" for c in db_cols])
            placeholders = ", ".join(["%s"] * len(db_cols))
            sql = f"INSERT INTO {config.ABD_TABLE_NAME} ({col_str}) VALUES ({placeholders})"

            # Iterate and insert rows
            rows_to_insert = []
            row_iterator = target_sheet_obj.iter_rows(min_row=2, values_only=True)
            for row in row_iterator:
                record = tuple(
                    row[excel_idx] if excel_idx < len(row) else None for db_col, excel_idx in col_map.items())
                rows_to_insert.append(record)

            if rows_to_insert:
                with tqdm(total=len(rows_to_insert), desc="    -> Inserting rows", unit="row") as pbar:
                    for record in rows_to_insert:
                        cursor.execute(sql, record)
                        pbar.update(1)
                connection.commit()
                print(f"  -> Successfully loaded {len(rows_to_insert)} records from {file_name}.")
            else:
                print("  -> No data rows found in this sheet.")

        except Exception as e:
            print(f"\n  -> ERROR: Could not process file {file_name}. Reason: {e}")

    print(f"\n✅ All {total_files} ABD files have been processed.")


# ▲▲▲ END OF REWRITTEN SECTION ▲▲▲


def import_pmr_data(connection, pmr_files):
    cursor = connection.cursor()
    pmr_df_list = [pd.read_excel(file) for file in pmr_files]
    df_all_pmr = pd.concat(pmr_df_list, ignore_index=True)
    df_all_pmr.columns = df_all_pmr.columns.str.strip().str.upper()

    for _, row in tqdm(df_all_pmr.iterrows(), total=len(df_all_pmr), desc="Loading PMR data      "):
        stripped_id = str(row.get('SAP PROJECT ID', '')).strip()
        final_project_id = stripped_id.lstrip('0') if stripped_id.isdigit() else stripped_id
        if final_project_id:
            manager_name = str(row.get('PROGRAM MANAGER NAME', '')).strip()
            manager_email = str(row.get('PROGRAM MANAGER EMAIL ID', '')).strip()
            sql = f"INSERT IGNORE INTO {config.PMR_TABLE} (PROJECT_ID, PGM_MANAGER_NAME, PGM_MANAGER_EMAIL) VALUES (%s, %s, %s)"
            cursor.execute(sql, (final_project_id, manager_name, manager_email))
    connection.commit()
    print("✅ PMR data loaded successfully (new entries added, existing entries ignored).")


def import_regional_details(connection, excel_path, fiscal_year):
    cursor = connection.cursor()
    cursor.execute(f"DELETE FROM {config.REGIONAL_TABLE} WHERE fiscal_year = %s", (fiscal_year,))

    xls = pd.ExcelFile(excel_path)
    sheet_name_pattern = re.compile(r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-\d{2}$")
    sheets_to_process = [s for s in xls.sheet_names if sheet_name_pattern.match(s)]

    for sheet_name in sheets_to_process:
        df = pd.read_excel(xls, sheet_name=sheet_name)
        df.columns = [col.strip().upper() for col in df.columns]
        agg_rules = {'TOTAL HOURS': 'sum', 'CURRENT WORK LOCATION': 'first', 'PROJECT DESCRIPTION': 'first',
                     'PROJECT TYPE': 'first', 'CONTRACT TYPE': 'first', 'CUST NAME': 'first', 'RUS STATUS': 'first'}
        df_agg = df.groupby(['EMPLID', 'PROJECT ID'], as_index=False).agg(agg_rules)
        parsed_date = datetime.strptime(sheet_name, '%b-%y')
        _, num_days = calendar.monthrange(parsed_date.year, parsed_date.month)
        end_of_month_date = parsed_date.replace(day=num_days).date()
        for _, row in tqdm(df_agg.iterrows(), total=len(df_agg), desc=f"Loading regional {sheet_name}", leave=False):
            sql = f"INSERT INTO {config.REGIONAL_TABLE} (fiscal_year, EMPLID, CURRENT_WORK_LOCATION, PROJECT_ID, PROJECT_DESCRIPTION, PROJECT_TYPE, CONTRACT_TYPE, CUST_NAME, RUS_STATUS, TOTAL_HOURS, Month) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
            values = (fiscal_year, str(row.get('EMPLID', '')).strip(),
                      str(row.get('CURRENT WORK LOCATION', '')).strip(), str(row.get('PROJECT ID', '')).strip(),
                      str(row.get('PROJECT DESCRIPTION', '')).strip(), str(row.get('PROJECT TYPE', '')).strip(),
                      str(row.get('CONTRACT TYPE', '')).strip(), str(row.get('CUST NAME', '')).strip(),
                      str(row.get('RUS STATUS', '')).strip(), row.get('TOTAL HOURS'), end_of_month_date)
            cursor.execute(sql, values)
    connection.commit()
    print(f"Regional data for {fiscal_year} loaded into {config.REGIONAL_TABLE}")


# --- MODIFIED: Handles the optional ER_NIC_SUM column ---
def import_salary_data(connection, excel_path, fiscal_year):
    cursor = connection.cursor()
    cursor.execute(f"DELETE FROM {config.SALARY_TABLE} WHERE fiscal_year = %s", (fiscal_year,))

    xls = pd.ExcelFile(excel_path)
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_name)
        df.columns = df.columns.str.strip().str.upper()

        # Check if the optional column exists
        has_er_nic_sum = 'ER_NIC_SUM' in df.columns

        for _, row in tqdm(df.iterrows(), total=len(df), desc=f"Loading salary {sheet_name: <15}", leave=False):
            month_date = pd.to_datetime(row['MONTH'])
            end_of_month_date = (month_date + pd.offsets.MonthEnd(0)).date()

            # Use .get() to safely retrieve the value, defaulting to None if it's not there
            er_nic_sum_value = row.get('ER_NIC_SUM') if has_er_nic_sum else None

            sql = f"INSERT INTO {config.SALARY_TABLE} (fiscal_year, EMPLID, MONTH, GROSS_PAY, ER_NIC_SUM) VALUES (%s, %s, %s, %s, %s)"
            values = (
                fiscal_year,
                str(row.get('EMPLID', '')).strip(),
                end_of_month_date,
                row.get('GROSS PAY'),
                er_nic_sum_value
            )
            cursor.execute(sql, values)

    connection.commit()
    print(f"Salary data for {fiscal_year} loaded into {config.SALARY_TABLE}")

def consolidate_data(connection, log_file, fiscal_year):
    """
    Consolidates regional, salary, and ABD data. It allocates monthly pay across
    projects based on hours worked to prevent duplication.
    """
    cursor = connection.cursor()
    print(f"Consolidating data for {fiscal_year}...")
    cursor.execute(f"DELETE FROM {config.CONSOLIDATION_TABLE} WHERE fiscal_year = %s", (fiscal_year,))

    # Step 1: Create a temporary table to hold allocated salary data
    cursor.execute(f"""
        CREATE TEMPORARY TABLE temp_allocated_salary (
            fiscal_year VARCHAR(10),
            EMPLID VARCHAR(255),
            Month DATE,
            GROSS_PAY DECIMAL(20, 2),
            ER_NIC_SUM DECIMAL(20, 2),
            PROJECT_ID VARCHAR(255)
        );
    """)

    # Step 2: Insert allocated salaries into the temporary table
    query_allocated = f"""
        INSERT INTO temp_allocated_salary
        SELECT
            s.fiscal_year, s.EMPLID, s.Month,
            s.GROSS_PAY * (r.TOTAL_HOURS / emh.total_hours),
            s.ER_NIC_SUM * (r.TOTAL_HOURS / emh.total_hours),
            r.PROJECT_ID
        FROM
            {config.SALARY_TABLE} s
        JOIN
            (SELECT EMPLID, Month, fiscal_year, SUM(TOTAL_HOURS) as total_hours
             FROM {config.REGIONAL_TABLE} WHERE fiscal_year = %s AND TOTAL_HOURS > 0
             GROUP BY EMPLID, Month, fiscal_year) emh
             ON s.EMPLID = emh.EMPLID AND s.Month = emh.Month AND s.fiscal_year = emh.fiscal_year
        JOIN
            {config.REGIONAL_TABLE} r ON emh.EMPLID = r.EMPLID AND emh.Month = r.Month AND emh.fiscal_year = r.fiscal_year
        WHERE
            s.fiscal_year = %s;
    """
    cursor.execute(query_allocated, (fiscal_year, fiscal_year))
    print(f"  - Calculated allocated pay for employees with project hours.")

    # Step 3: Insert unallocated salaries into the temporary table
    query_unallocated = f"""
        INSERT INTO temp_allocated_salary
        SELECT
            s.fiscal_year, s.EMPLID, s.Month, s.GROSS_PAY, s.ER_NIC_SUM, NULL
        FROM
            {config.SALARY_TABLE} s
        LEFT JOIN
            (SELECT DISTINCT EMPLID, Month, fiscal_year FROM {config.REGIONAL_TABLE}
             WHERE fiscal_year = %s AND TOTAL_HOURS > 0) r_valid
             ON s.EMPLID = r_valid.EMPLID AND s.Month = r_valid.Month AND s.fiscal_year = r_valid.fiscal_year
        WHERE
            s.fiscal_year = %s AND r_valid.EMPLID IS NULL;
    """
    cursor.execute(query_unallocated, (fiscal_year, fiscal_year))
    print(f"  - Added pay for salary-only employees.")

    # Step 4: Insert the final consolidated data into the main table
    final_insert_query = f"""
        INSERT INTO {config.CONSOLIDATION_TABLE} (
            fiscal_year, EMPLID, Month, GROSS_PAY, ER_NIC_SUM,
            DESIGNATION, BAND, `FUNCTION`,
            CURRENT_WORK_LOCATION, PROJECT_ID, PROJECT_DESCRIPTION, PROJECT_TYPE,
            CONTRACT_TYPE, CUST_NAME, PGM_MANAGER_NAME, PGM_MANAGER_EMAIL
        )
        SELECT
            t.fiscal_year, t.EMPLID, t.Month, t.GROSS_PAY, t.ER_NIC_SUM,
            abd.Designation, abd.BAND, abd.Function,
            r.CURRENT_WORK_LOCATION, r.PROJECT_ID, r.PROJECT_DESCRIPTION,
            r.PROJECT_TYPE, r.CONTRACT_TYPE, r.CUST_NAME,
            COALESCE(pmr.PGM_MANAGER_NAME, abd.PROGRAM_MANAGER_NAME),
            pmr.PGM_MANAGER_EMAIL
        FROM
            temp_allocated_salary t
        LEFT JOIN
            {config.REGIONAL_TABLE} r ON t.EMPLID = r.EMPLID AND t.Month = r.Month AND t.fiscal_year = r.fiscal_year AND t.PROJECT_ID = r.PROJECT_ID
        LEFT JOIN
            (SELECT * FROM {config.ABD_DB_NAME}.{config.ABD_TABLE_NAME}
             WHERE (EMPLID, id) IN
                (SELECT EMPLID, MAX(id) FROM {config.ABD_DB_NAME}.{config.ABD_TABLE_NAME} GROUP BY EMPLID)
            ) abd ON t.EMPLID = abd.EMPLID
        LEFT JOIN
            {config.PMR_DB_NAME}.{config.PMR_TABLE} pmr ON t.PROJECT_ID = pmr.PROJECT_ID;
    """
    cursor.execute(final_insert_query)
    print(f"  - Final consolidation complete.")

    # Step 5: Drop the temporary table
    cursor.execute("DROP TEMPORARY TABLE temp_allocated_salary;")

    connection.commit()
    print(f"Data for {fiscal_year} consolidated successfully.")

    # --- The rest of the function (logging missing projects) remains the same ---
    missing_query = f"""
        SELECT DISTINCT r.PROJECT_ID
        FROM {config.REGIONAL_TABLE} r
        LEFT JOIN {config.PMR_DB_NAME}.{config.PMR_TABLE} pmr ON r.PROJECT_ID = pmr.PROJECT_ID
        WHERE r.fiscal_year = %s AND pmr.PROJECT_ID IS NULL 
        AND r.PROJECT_ID IS NOT NULL AND r.PROJECT_ID != ''
    """
    cursor.execute(missing_query, (fiscal_year,))
    missing_projects = [row[0] for row in cursor.fetchall()]
    with open(log_file, "w") as log:
        log.write(f"Missing Project IDs for {fiscal_year} (not found in PMR table):\n")
        if missing_projects:
            log.write("\n".join(sorted(missing_projects)))
    print(f"Missing projects for {fiscal_year} logged in {log_file}.")

def fill_missing_emails(connection, db_name, fiscal_year):
    """
    Updates the consolidation table to fill in missing PGM emails by
    matching the PGM name against the global PMR data.
    """
    print(f"Attempting to fill missing PGM emails for {fiscal_year} by name matching...")
    cursor = connection.cursor()

    # ▼▼▼ THIS QUERY HAS BEEN UPDATED ▼▼▼
    # This query joins the consolidation table with the global PMR table
    # on the cleaned manager name (case-insensitive and trimmed)
    update_query = f"""
        UPDATE
            {db_name}.{config.CONSOLIDATION_TABLE} c
        JOIN
            (SELECT 
                 PGM_MANAGER_NAME, 
                 MAX(PGM_MANAGER_EMAIL) AS PGM_MANAGER_EMAIL 
             FROM {config.PMR_DB_NAME}.{config.PMR_TABLE}
             WHERE PGM_MANAGER_EMAIL IS NOT NULL AND PGM_MANAGER_NAME IS NOT NULL
             GROUP BY PGM_MANAGER_NAME
            ) AS pmr_unique 
            ON LOWER(TRIM(c.PGM_MANAGER_NAME)) = LOWER(TRIM(pmr_unique.PGM_MANAGER_NAME))
        SET
            c.PGM_MANAGER_EMAIL = pmr_unique.PGM_MANAGER_EMAIL
        WHERE
            c.PGM_MANAGER_EMAIL IS NULL
            AND c.fiscal_year = %s;
    """
    # ▲▲▲ END OF UPDATED SECTION ▲▲▲

    try:
        cursor.execute(update_query, (fiscal_year,))
        connection.commit()
        # cursor.rowcount provides the number of rows updated
        print(f"  -> Success: {cursor.rowcount} missing emails were found and filled.")
    except Error as e:
        print(f"  -> An error occurred during email backfill: {e}")